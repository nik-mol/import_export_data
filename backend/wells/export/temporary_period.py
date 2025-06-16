from datetime import date

import numpy as np
from common.db_connector import ConnectorManager
from common.excel_templates.custom_worksheet import CustomWorkSheet
from dateutil.relativedelta import relativedelta
from django.db.models import Max
from openpyxl import Workbook
from wells.excel_templates.config.description_sheets import TemporaryPeriodHeadersConfig
from wells.excel_templates.config.temporary_period import WorkSheetHeaderTemporaryPeriod
from wells.excel_templates.temporary_period import TemporaryPeriodTemplate
from wells.export.config import SheetExcelData
from wells.models import WellsBaseFund
from wells.queries import CalculationTemporaryPeriodLoader


class ExportCalculationTemporaryPeriod:
    """
    Класс для выгрузки отчета по временным приостановкам из ФОНДа ИНК
    """

    @staticmethod
    def _write_data_in_excel(input_date: date) -> Workbook:
        """
        Запись данных в эксель

        Returns:
                Workbook: файл электронной таблицы
        """

        workbook: Workbook = TemporaryPeriodTemplate()

        results: dict[float, np.ndarray] = {}
        for sheet_number in [1, 2, 3, 4.1, 4.2]:
            results[sheet_number] = ConnectorManager.get_raw_data(
                CalculationTemporaryPeriodLoader.get_wells(sheet_number)
            )

        wells_output_temporary_period = ConnectorManager.get_raw_data(
            CalculationTemporaryPeriodLoader.get_wells_output_temporary_period(
                input_date
            )
        )

        sheet_info = [
            (results[1], "Фонд"),
            (results[2], "вр_приост_1"),
            (results[3], "вр_приост_продление"),
            (results[4.1], "вр_приост_2"),
            (results[4.2], "вр_приост_2"),
            (wells_output_temporary_period, "вывод_из_вр_приост"),
        ]
        result_wells: list[SheetExcelData] = [
            SheetExcelData(wells_loader=loader, sheet_name=name)
            for loader, name in sheet_info
        ]

        last_date = ExportCalculationTemporaryPeriod.__get_last_date()
        for wells in result_wells:
            ExportCalculationTemporaryPeriod.__write_data_in_sheet(
                wells.wells_loader, workbook[wells.sheet_name]
            )
            ExportCalculationTemporaryPeriod.__write_headers_with_date_sheet(
                workbook[wells.sheet_name], last_date, input_date
            )

        return workbook

    @staticmethod
    def __write_data_in_sheet(
        wells: np.ndarray, active_worksheet: CustomWorkSheet
    ) -> None:
        """
        Запись данных во вкладку экселя

        Args:
        wells (np.ndarray): данные для записи в эксель
        active_worksheet (CustomWorkSheet): оптимизированная рабочая книга
        """

        date_letters = [
            WorkSheetHeaderTemporaryPeriod.BUILDING_END_DATE.value[0]["column"],
            WorkSheetHeaderTemporaryPeriod.DELAY_PERIOD.value[0]["column"],
            WorkSheetHeaderTemporaryPeriod.DELAY_START.value[0]["column"],
        ]

        if len(wells):
            for item in wells:
                active_worksheet.append(list(item))

                if active_worksheet.title != "вывод_из_вр_приост":
                    for date_letter in date_letters:
                        cell = active_worksheet.cell(
                            column=date_letter,
                            row=active_worksheet._max_row,
                        )
                        cell.number_format = "dd.mm.yyyy"

    @staticmethod
    def __write_headers_with_date_sheet(
        active_worksheet: CustomWorkSheet, last_date: date, input_date: date
    ) -> None:
        """
        Заполнение наименований колонок в формате дат

        Args:
            active_worksheet (CustomWorkSheet): оптимизированная рабочая книга
            last_date (date): крайняя дата фонда ИНК
            input_date (date): дата для формиорвания листа "вывод_из_вр_приост"
        """

        if active_worksheet.title != "вывод_из_вр_приост":
            column = TemporaryPeriodHeadersConfig.COLUMN_LAST_DATE_SHEETS_1_4.value
        else:
            column = TemporaryPeriodHeadersConfig.COLUMN_LAST_DATE_SHEETS_5.value
            if input_date:
                last_date = input_date
            if last_date:
                previous_date = last_date - relativedelta(months=1)
                active_worksheet.cell(
                    row=TemporaryPeriodHeadersConfig.ROW.value,
                    column=TemporaryPeriodHeadersConfig.COLUMN_PREVIOUS_DATE_SHEETS_5.value,
                    value=f'Статус на {previous_date.strftime("%d.%m.%Y")}',
                )

        if last_date:
            active_worksheet.cell(
                row=TemporaryPeriodHeadersConfig.ROW.value,
                column=column,
                value=f'Статус на {last_date.strftime("%d.%m.%Y")}',
            )

    @staticmethod
    def __get_last_date() -> date:
        """
        Крайняя дата

        Returns
           date: Крайняя дата фонда ИНК
        """

        return WellsBaseFund.objects.aggregate(Max("date")).get("date__max")
