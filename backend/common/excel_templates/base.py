from dataclasses import dataclass
from enum import Enum
from typing import Dict, List, Optional, Union

from common.excel_templates.custom_worksheet import CustomWorkSheet as Worksheet
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass
class ExcelCellFontStyle:
    """
    Параметры текста для ячейки
    """

    name: str
    size: str
    color: str


@dataclass
class ExcelCellAligment:
    """
    Положение ячейки
    """

    horizontal: Optional[str] = None
    vertical: Optional[str] = None
    wrap_text: bool = True


@dataclass
class ExcelCellStyle:
    """
    Параметры самой ячейки
    """

    aligment: Optional[ExcelCellAligment] = None
    fill: Optional[Dict] = None
    border: Optional[Dict] = None


@dataclass
class ExcelTemplateElement:
    """
    Элемент подписи данных для шаблона экселя
    """

    column: int
    row: int
    text: str
    font_style: ExcelCellFontStyle
    cell_style: Optional[Union[ExcelCellStyle | dict]] = None
    column_width: Optional[int] = None
    merge_cell: Optional[bool] = False
    column_end: Optional[int] = None
    row_end: Optional[int] = None


@dataclass
class ExcelCellFormatStyle:
    """
    Датакласс для стилизации ячейки
    """

    font_style: ExcelCellFontStyle
    cell_style: Optional[ExcelCellStyle] = None
    column_width: Optional[int] = None


class ExcelWorkSheet:
    """
    Страница в экселе
    """

    header: type[Enum]
    title: str
    worksheet_seal: List[dict]

    def __new__(cls, workbook: Workbook, new_attrs: Optional[dict] = None) -> Worksheet:
        """
        Метод возращает инстанс листа
        """
        if new_attrs is not None:
            for key, value in new_attrs.items():
                setattr(cls, key, value)

        new_worksheet = Worksheet(parent=workbook, title=cls.title)
        workbook._add_sheet(sheet=new_worksheet)

        try:
            cls.__create_table_header(new_worksheet)
        except AttributeError:
            pass

        return new_worksheet

    @classmethod
    def __create_table_header(cls, worksheet: Worksheet):
        """
        Метод создает шапку на переданном листе
        """
        for data in cls.header:
            cell_element = ExcelTemplateElement(
                **data.value[0]
            )  # Когда перебираем энум, он дает tuples
            cls.set_worksheet_cell_value(worksheet, cell_element)

    @staticmethod
    def set_worksheet_cell_value(worksheet: Worksheet, cell: ExcelTemplateElement):
        """
        Установка параметров определенной ячейке
        """
        if cell.merge_cell is True:
            worksheet.merge_cells(
                start_row=cell.row,
                start_column=cell.column,
                end_row=cell.row_end,
                end_column=cell.column_end,
            )
        created_cell = worksheet.cell(row=cell.row, column=cell.column)
        created_cell.value = cell.text
        if cell.font_style:
            created_cell.font = Font(**cell.font_style)

        if cell.column_width:
            column_string = get_column_letter(cell.column)
            worksheet.column_dimensions[column_string].width = cell.column_width
        if isinstance(cell.cell_style, dict) and cell.cell_style:
            cell.cell_style = ExcelCellStyle(**cell.cell_style)
            if cell.cell_style.aligment:
                created_cell.alignment = Alignment(**cell.cell_style.aligment)
            if cell.cell_style.fill:
                created_cell.fill = PatternFill(**cell.cell_style.fill)
            if cell.cell_style.border:
                cell.cell_style.border = ExcelNamedStyleCreator.create_borders(
                    cell.cell_style.border
                )
                created_cell.border = Border(**cell.cell_style.border)


class ExcelNamedStyleCreator:
    """
    Создание именованого стиля для ячеек экселя
    """

    @staticmethod
    def create_named_style(
        cell: ExcelCellFormatStyle, name: str, round_by=None, builtinId=None
    ) -> NamedStyle:
        """
        Создание именованного стиля из параметров ячейки

        Args:
            cell (ExcelTemplateElement): _description_
            name (str): _description_

        Returns:
            NamedStyle: _description_
        """
        if builtinId is None:
            named_style = NamedStyle(name=name)
        else:
            named_style = NamedStyle(name=name, builtinId=builtinId)
        if round_by is not None:
            named_style.number_format = "0." + "0" * round_by if round_by != 0 else "0"
        if cell.font_style:
            named_style.font = Font(**cell.font_style)
        if isinstance(cell.cell_style, dict) and cell.cell_style:
            cell.cell_style = ExcelCellStyle(**cell.cell_style)
            if cell.cell_style.aligment:
                named_style.alignment = Alignment(**cell.cell_style.aligment)
            if cell.cell_style.fill:
                named_style.fill = PatternFill(**cell.cell_style.fill)
            if cell.cell_style.border:
                cell.cell_style.border = ExcelNamedStyleCreator.create_borders(
                    cell.cell_style.border
                )
                named_style.border = Border(**cell.cell_style.border)
        return named_style

    @staticmethod
    def create_borders(borders_style_dict: Dict):
        """
        Метод создает границы для ячейки
        """
        output_dict = {}
        for key, value in borders_style_dict.items():
            output_dict[key] = Side(**value)
        return output_dict


class ExcelTemplate:
    """
    Шаблон формы экселя
    """

    title: str
    worksheet_list: List[type[ExcelWorkSheet]]

    def __new__(cls) -> Workbook:
        new_workbook = Workbook(write_only=True)
        for worksheet in cls.worksheet_list:
            worksheet(new_workbook)
        return new_workbook
